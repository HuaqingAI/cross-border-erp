from __future__ import annotations

import asyncio
from io import BytesIO
import uuid

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.import_task import ImportTask
from app.models.system_enum import SystemEnum
from app.models.user import User, UserRole


async def _login_as_role(
    client: AsyncClient,
    db_session: AsyncSession,
    role: UserRole,
) -> None:
    username = f"{role.value}-{uuid.uuid4().hex[:8]}"
    user = User(
        username=username,
        password_hash=hash_password("Test123!"),
        role=role,
        is_active=True,
    )
    db_session.add(user)
    await db_session.commit()

    response = await client.post(
        "/api/v1/auth/login",
        json={"username": username, "password": "Test123!"},
    )
    assert response.status_code == 200


async def _seed_enum(
    db_session: AsyncSession,
    *,
    enum_group: str,
    enum_key: str,
    enum_value: str | None = None,
) -> None:
    existing = await db_session.scalar(
        select(SystemEnum).where(
            SystemEnum.enum_group == enum_group,
            SystemEnum.enum_key == enum_key,
            SystemEnum.deleted_at.is_(None),
        )
    )
    if existing is not None:
        existing.is_enabled = True
        existing.enum_value = enum_value or enum_key
        db_session.add(existing)
        await db_session.commit()
        return

    db_session.add(
        SystemEnum(
            enum_group=enum_group,
            enum_key=enum_key,
            enum_value=enum_value or enum_key,
            is_enabled=True,
            sort_order=10,
        )
    )
    await db_session.commit()


async def _seed_sku_enums(db_session: AsyncSession) -> None:
    await _seed_enum(db_session, enum_group="product_type", enum_key="主品")
    await _seed_enum(db_session, enum_group="product_status", enum_key="上架")


def _build_excel(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    instruction = workbook.active
    instruction.title = "填写说明"
    instruction.append(["说明", "测试文件"])
    data_sheet = workbook.create_sheet("数据")
    data_sheet.append(headers)
    for row in rows:
        data_sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def _create_category(
    client: AsyncClient,
    *,
    code: str,
    name: str,
    parent_id: int | None = None,
) -> int:
    response = await client.post(
        "/api/v1/products/categories",
        json={"code": code, "name": name, "parent_id": parent_id},
    )
    assert response.status_code == 201
    return response.json()["id"]


async def _create_spu_with_supplier(
    client: AsyncClient,
    *,
    code: str,
    supplier_name: str,
) -> None:
    level1_id = await _create_category(client, code=f"{code}-L1", name=f"{code}-一级")
    level2_id = await _create_category(client, code=f"{code}-L2", name=f"{code}-二级", parent_id=level1_id)
    level3_id = await _create_category(client, code=f"{code}-L3", name=f"{code}-三级", parent_id=level2_id)
    response = await client.post(
        "/api/v1/spus",
        json={
            "code": code,
            "name": f"{code}-名称",
            "level1_category_id": level1_id,
            "level2_category_id": level2_id,
            "level3_category_id": level3_id,
            "customer_warranty_months": 24,
            "unit": "台",
            "restricted_countries": ["US"],
            "supplier_name": supplier_name,
            "manufacturer_model": "MODEL-A",
            "purchase_price": "88.00",
            "purchase_warranty_months": 12,
            "supplier_warranty_notes": "说明",
            "invoice_infos": [
                {
                    "invoice_name": "开票名",
                    "invoice_unit": "台",
                    "invoice_model": "INV-1",
                    "company_subject": "抬头A",
                    "sort_order": 0,
                }
            ],
        },
    )
    assert response.status_code == 201


@pytest.mark.asyncio
async def test_download_category_template_returns_excel(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)

    response = await client.get("/api/v1/import/templates/categories")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["填写说明", "数据"]
    assert workbook["数据"]["A1"].value == "分类编码"


@pytest.mark.asyncio
async def test_business_user_cannot_use_import_apis(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.BUSINESS_DEPT)

    response = await client.get("/api/v1/import/templates/categories")

    assert response.status_code == 403


@pytest.mark.asyncio
async def test_category_import_validates_and_confirms(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    file_bytes = _build_excel(
        ["分类编码", "分类名称", "层级", "父级分类编码", "排序"],
        [
            ["C-L1", "一级分类", 1, "", 10],
            ["C-L2", "二级分类", 2, "C-L1", 20],
            ["C-L3", "三级分类", 3, "C-L2", 30],
        ],
    )

    validate_response = await client.post(
        "/api/v1/import/categories",
        files={"file": ("categories.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert validate_response.status_code == 201
    data = validate_response.json()
    assert data["status"] == "validated"
    assert data["success_count"] == 3
    assert data["failed_count"] == 0
    assert data["can_confirm"] is True

    confirm_response = await client.post(
        "/api/v1/import/categories/confirm",
        json={"task_id": data["task_id"]},
    )

    assert confirm_response.status_code == 200
    assert confirm_response.json()["status"] == "imported"
    tree_response = await client.get("/api/v1/products/categories/tree")
    tree = tree_response.json()
    assert len(tree) == 1
    assert tree[0]["code"] == "C-L1"
    assert tree[0]["children"][0]["code"] == "C-L2"
    assert tree[0]["children"][0]["children"][0]["code"] == "C-L3"


@pytest.mark.asyncio
async def test_category_import_overlong_fields_return_row_errors_instead_of_500(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    file_bytes = _build_excel(
        ["分类编码", "分类名称", "层级", "父级分类编码", "排序"],
        [
            ["C" * 60, "超长分类名称", 1, "", 10],
        ],
    )

    response = await client.post(
        "/api/v1/import/categories",
        files={"file": ("categories-long.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["status"] == "failed_validation"
    assert data["success_count"] == 0
    assert data["failed_count"] == 1
    assert any(item["field"] == "分类编码" for item in data["errors"])


@pytest.mark.asyncio
async def test_spu_import_reports_invalid_category_reference(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    await _create_spu_with_supplier(client, code="SPU-SUPPLIER-SEED-A", supplier_name="供应商A")
    file_bytes = _build_excel(
        [
            "SPU编码",
            "SPU名称",
            "一级分类编码",
            "二级分类编码",
            "三级分类编码",
            "客户质保期(月)",
            "单位",
            "禁止经营国家编码",
            "供应商名称",
            "厂家型号",
            "采购价",
            "采购质保期(月)",
            "供应商质保说明",
            "开票品名",
            "开票单位",
            "开票型号",
            "公司抬头",
            "开票排序",
        ],
        [
            ["SPU-IMPORT-001", "导入SPU", "L1-NOT-FOUND", "L2", "L3", 24, "台", "US,DE", "供应商A", "M-1", "99.50", 12, "说明", "开票名", "台", "INV", "抬头", 1],
        ],
    )

    response = await client.post(
        "/api/v1/import/spus",
        files={"file": ("spus.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["status"] == "failed_validation"
    assert data["failed_count"] > 0
    assert data["can_confirm"] is False
    assert any(item["field"] == "一级分类编码" for item in data["errors"])


@pytest.mark.asyncio
async def test_spu_import_mixed_valid_and_invalid_records_report_both_counts(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    await _create_spu_with_supplier(client, code="SPU-SUPPLIER-SEED-MIX", supplier_name="供应商A")
    level1_id = await _create_category(client, code="MIX-L1", name="一级")
    level2_id = await _create_category(client, code="MIX-L2", name="二级", parent_id=level1_id)
    await _create_category(client, code="MIX-L3", name="三级", parent_id=level2_id)

    file_bytes = _build_excel(
        [
            "SPU编码",
            "SPU名称",
            "一级分类编码",
            "二级分类编码",
            "三级分类编码",
            "客户质保期(月)",
            "单位",
            "禁止经营国家编码",
            "供应商名称",
            "厂家型号",
            "采购价",
            "采购质保期(月)",
            "供应商质保说明",
            "开票品名",
            "开票单位",
            "开票型号",
            "公司抬头",
            "开票排序",
        ],
        [
            ["SPU-MIX-OK", "有效SPU", "MIX-L1", "MIX-L2", "MIX-L3", 24, "台", "US", "供应商A", "M-1", "99.50", 12, "说明", "开票名", "台", "INV", "抬头", 1],
            ["SPU-MIX-BAD", "无效SPU", "MIX-L1", "MIX-L2", "MIX-L3", 24, "台", "US", "全新供应商", "M-2", "99.50", 12, "说明", "开票名", "台", "INV", "抬头", 1],
        ],
    )

    response = await client.post(
        "/api/v1/import/spus",
        files={"file": ("spus-mixed.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["status"] == "failed_validation"
    assert data["success_count"] == 1
    assert data["failed_count"] == 1
    assert any(item["row_key"] == "SPU-MIX-BAD" for item in data["errors"])


@pytest.mark.asyncio
async def test_spu_import_can_group_invoice_rows_and_confirm(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    await _create_spu_with_supplier(client, code="SPU-SUPPLIER-SEED-B", supplier_name="供应商A")
    level1_id = await _create_category(client, code="SPU-L1", name="一级")
    level2_id = await _create_category(client, code="SPU-L2", name="二级", parent_id=level1_id)
    await _create_category(client, code="SPU-L3", name="三级", parent_id=level2_id)

    file_bytes = _build_excel(
        [
            "SPU编码",
            "SPU名称",
            "一级分类编码",
            "二级分类编码",
            "三级分类编码",
            "客户质保期(月)",
            "单位",
            "禁止经营国家编码",
            "供应商名称",
            "厂家型号",
            "采购价",
            "采购质保期(月)",
            "供应商质保说明",
            "开票品名",
            "开票单位",
            "开票型号",
            "公司抬头",
            "开票排序",
        ],
        [
            ["SPU-IMPORT-OK", "导入SPU", "SPU-L1", "SPU-L2", "SPU-L3", 24, "台", "US,DE", "供应商A", "M-1", "99.50", 12, "说明", "开票名1", "台", "INV-1", "抬头A", 1],
            ["SPU-IMPORT-OK", "导入SPU", "SPU-L1", "SPU-L2", "SPU-L3", 24, "台", "US,DE", "供应商A", "M-1", "99.50", 12, "说明", "开票名2", "台", "INV-2", "抬头A", 2],
        ],
    )

    validate_response = await client.post(
        "/api/v1/import/spus",
        files={"file": ("spus.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    task_id = validate_response.json()["task_id"]

    confirm_response = await client.post(
        "/api/v1/import/spus/confirm",
        json={"task_id": task_id},
    )

    assert confirm_response.status_code == 200
    spu_response = await client.get("/api/v1/spus", params={"keyword": "SPU-IMPORT-OK"})
    spu_data = spu_response.json()
    assert spu_data["total"] == 1
    detail_response = await client.get(f"/api/v1/spus/{spu_data['items'][0]['id']}")
    assert len(detail_response.json()["invoice_infos"]) == 2


@pytest.mark.asyncio
async def test_spu_import_rejects_supplier_not_existing_in_system(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    level1_id = await _create_category(client, code="SUP-L1", name="一级")
    level2_id = await _create_category(client, code="SUP-L2", name="二级", parent_id=level1_id)
    await _create_category(client, code="SUP-L3", name="三级", parent_id=level2_id)

    file_bytes = _build_excel(
        [
            "SPU编码",
            "SPU名称",
            "一级分类编码",
            "二级分类编码",
            "三级分类编码",
            "客户质保期(月)",
            "单位",
            "禁止经营国家编码",
            "供应商名称",
            "厂家型号",
            "采购价",
            "采购质保期(月)",
            "供应商质保说明",
            "开票品名",
            "开票单位",
            "开票型号",
            "公司抬头",
            "开票排序",
        ],
        [
            ["SPU-NO-SUPPLIER", "导入SPU", "SUP-L1", "SUP-L2", "SUP-L3", 24, "台", "US", "全新供应商", "M-1", "99.50", 12, "说明", "开票名", "台", "INV", "抬头", 1],
        ],
    )

    response = await client.post(
        "/api/v1/import/spus",
        files={"file": ("spus.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["status"] == "failed_validation"
    assert any(item["field"] == "供应商名称" and item["message"] == "供应商必须已存在" for item in data["errors"])


@pytest.mark.asyncio
async def test_spu_import_failed_count_counts_records_not_error_items(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    await _create_spu_with_supplier(client, code="SPU-SUPPLIER-SEED-C", supplier_name="供应商A")
    level1_id = await _create_category(client, code="COUNT-L1", name="一级")
    level2_id = await _create_category(client, code="COUNT-L2", name="二级", parent_id=level1_id)
    await _create_category(client, code="COUNT-L3", name="三级", parent_id=level2_id)

    file_bytes = _build_excel(
        [
            "SPU编码",
            "SPU名称",
            "一级分类编码",
            "二级分类编码",
            "三级分类编码",
            "客户质保期(月)",
            "单位",
            "禁止经营国家编码",
            "供应商名称",
            "厂家型号",
            "采购价",
            "采购质保期(月)",
            "供应商质保说明",
            "开票品名",
            "开票单位",
            "开票型号",
            "公司抬头",
            "开票排序",
        ],
        [
            ["SPU-COUNT-FAIL", "导入SPU", "COUNT-L1", "COUNT-L2", "COUNT-L3", 24, "台", "US", "供应商A", "M-1", "99.50", 12, "说明", "", "", "INV", "抬头", 1],
        ],
    )

    response = await client.post(
        "/api/v1/import/spus",
        files={"file": ("spus.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["failed_count"] == 1
    assert len(data["errors"]) >= 2


@pytest.mark.asyncio
async def test_category_confirm_exposes_importing_progress_during_execution(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    file_bytes = _build_excel(
        ["分类编码", "分类名称", "层级", "父级分类编码", "排序"],
        [
            ["PROGRESS-L1", "一级分类", 1, "", 10],
            ["PROGRESS-L2", "二级分类", 2, "PROGRESS-L1", 20],
            ["PROGRESS-L3", "三级分类", 3, "PROGRESS-L2", 30],
        ],
    )
    validate_response = await client.post(
        "/api/v1/import/categories",
        files={"file": ("categories.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    task_id = validate_response.json()["task_id"]

    from app.services.product_categories import ProductCategoryService

    original_create = ProductCategoryService.create_category

    async def delayed_create(self, data):
        await asyncio.sleep(0.05)
        return await original_create(self, data)

    monkeypatch.setattr(ProductCategoryService, "create_category", delayed_create)

    confirm_task = asyncio.create_task(
        client.post("/api/v1/import/categories/confirm", json={"task_id": task_id})
    )
    await asyncio.sleep(0.02)
    progress_response = await client.get(f"/api/v1/import/tasks/{task_id}")
    confirm_response = await confirm_task

    assert progress_response.status_code == 200
    progress_data = progress_response.json()
    assert progress_data["status"] == "importing"
    assert progress_data["progress_percent"] < 100
    assert confirm_response.status_code == 200


@pytest.mark.asyncio
async def test_sku_import_validates_and_confirms_with_customs_payload(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    await _seed_sku_enums(db_session)
    level1_id = await _create_category(client, code="SKU-L1", name="一级")
    level2_id = await _create_category(client, code="SKU-L2", name="二级", parent_id=level1_id)
    level3_id = await _create_category(client, code="SKU-L3", name="三级", parent_id=level2_id)
    spu_payload = {
        "code": "SPU-FOR-SKU-IMPORT",
        "name": "被导入SPU",
        "level1_category_id": level1_id,
        "level2_category_id": level2_id,
        "level3_category_id": level3_id,
        "customer_warranty_months": 24,
        "unit": "台",
        "restricted_countries": ["US"],
        "supplier_name": "供应商A",
        "manufacturer_model": "MODEL-A",
        "purchase_price": "88.00",
        "purchase_warranty_months": 12,
        "supplier_warranty_notes": "说明",
        "invoice_infos": [
            {
                "invoice_name": "开票名",
                "invoice_unit": "台",
                "invoice_model": "INV-1",
                "company_subject": "抬头A",
                "sort_order": 0,
            }
        ],
    }
    create_spu_response = await client.post("/api/v1/spus", json=spu_payload)
    assert create_spu_response.status_code == 201

    file_bytes = _build_excel(
        [
            "所属SPU编码",
            "SKU编码",
            "SKU中文名称",
            "SKU英文名称",
            "产品型号",
            "产品类型",
            "核心参数",
            "产品状态",
            "电气参数",
            "产品原理",
            "用途",
            "材质",
            "单位",
            "是否带插头",
            "是否特殊产品",
            "特殊说明",
            "包装类型",
            "包装数量",
            "海关HS编码",
            "海关监管条件",
            "报关要素",
            "退税率",
            "报关信息已就绪",
            "包装净重KG",
            "包装毛重KG",
            "包装长CM",
            "包装宽CM",
            "包装高CM",
            "包装体积CBM",
            "包装排序",
        ],
        [
            ["SPU-FOR-SKU-IMPORT", "SKU-IMPORT-001", "导入SKU", "Imported SKU", "MODEL-1", "主品", "核心参数", "上架", "220V/50Hz", "原理", "用途", "ABS", "台", "是", "否", "", "纸箱", 2, "9018", "A", "要素A", "13", "是", "1.2", "1.5", "10", "20", "30", "0.006", 0],
        ],
    )

    validate_response = await client.post(
        "/api/v1/import/skus",
        files={"file": ("skus.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert validate_response.status_code == 201
    task_id = validate_response.json()["task_id"]

    confirm_response = await client.post(
        "/api/v1/import/skus/confirm",
        json={"task_id": task_id},
    )
    assert confirm_response.status_code == 200

    list_response = await client.get("/api/v1/skus", params={"keyword": "SKU-IMPORT-001"})
    list_data = list_response.json()
    assert list_data["total"] == 1

    detail_response = await client.get(f"/api/v1/skus/{list_data['items'][0]['id']}")
    detail = detail_response.json()
    assert detail["customs_hscode"] == "9018"
    assert detail["customs_info_ready"] is True
    assert detail["package_details"][0]["gross_weight_kg"] == "1.500"

    task = await db_session.scalar(select(ImportTask).where(ImportTask.id == task_id))
    assert task is not None
    assert task.status == "imported"


@pytest.mark.asyncio
async def test_sku_import_overlong_fields_return_row_errors_instead_of_500(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await _login_as_role(client, db_session, UserRole.ADMIN)
    await _seed_sku_enums(db_session)
    level1_id = await _create_category(client, code="SKU-LONG-L1", name="一级")
    level2_id = await _create_category(client, code="SKU-LONG-L2", name="二级", parent_id=level1_id)
    level3_id = await _create_category(client, code="SKU-LONG-L3", name="三级", parent_id=level2_id)
    create_spu_response = await client.post(
        "/api/v1/spus",
        json={
            "code": "SPU-FOR-SKU-LONG",
            "name": "被导入SPU",
            "level1_category_id": level1_id,
            "level2_category_id": level2_id,
            "level3_category_id": level3_id,
            "customer_warranty_months": 24,
            "unit": "台",
            "restricted_countries": ["US"],
            "supplier_name": "供应商A",
            "manufacturer_model": "MODEL-A",
            "purchase_price": "88.00",
            "purchase_warranty_months": 12,
            "supplier_warranty_notes": "说明",
            "invoice_infos": [
                {
                    "invoice_name": "开票名",
                    "invoice_unit": "台",
                    "invoice_model": "INV-1",
                    "company_subject": "抬头A",
                    "sort_order": 0,
                }
            ],
        },
    )
    assert create_spu_response.status_code == 201

    file_bytes = _build_excel(
        [
            "所属SPU编码",
            "SKU编码",
            "SKU中文名称",
            "SKU英文名称",
            "产品型号",
            "产品类型",
            "核心参数",
            "产品状态",
            "电气参数",
            "产品原理",
            "用途",
            "材质",
            "单位",
            "是否带插头",
            "是否特殊产品",
            "特殊说明",
            "包装类型",
            "包装数量",
            "海关HS编码",
            "海关监管条件",
            "报关要素",
            "退税率",
            "报关信息已就绪",
            "包装净重KG",
            "包装毛重KG",
            "包装长CM",
            "包装宽CM",
            "包装高CM",
            "包装体积CBM",
            "包装排序",
        ],
        [
            ["SPU-FOR-SKU-LONG", "SKU-LONG-001", "导入SKU", "Imported SKU", "M" * 101, "主品", "核心参数", "上架", "220V/50Hz", "原理", "用途", "ABS", "台", "是", "否", "", "纸箱", 2, "9018", "A", "要素A", "13", "是", "1.2", "1.5", "10", "20", "30", "0.006", 0],
        ],
    )

    response = await client.post(
        "/api/v1/import/skus",
        files={"file": ("skus-long.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["status"] == "failed_validation"
    assert data["success_count"] == 0
    assert data["failed_count"] == 1
    assert any(item["field"] == "产品型号" for item in data["errors"])
