from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.exceptions import BusinessError
from app.models.import_task import ImportTask
from app.models.product_category import ProductCategory
from app.models.spu import SPU
from app.models.sku import SKU
from app.models.system_enum import SystemEnum
from app.models.user import User
from app.repositories.import_tasks import ImportTaskRepository
from app.repositories.product_categories import ProductCategoryRepository
from app.repositories.spus import SPURepository
from app.repositories.skus import SKURepository
from app.schemas.import_tasks import (
    ImportConfirmResponse,
    ImportTaskProgressResponse,
    ImportValidationErrorItem,
    ImportValidationResponse,
)
from app.schemas.product_category import ProductCategoryCreate
from app.schemas.spu import SPUCreate, SPUInvoiceInfoPayload
from app.schemas.sku import (
    SKUCreate,
    SKUCustomsInfoUpdate,
    SKUPackageDetailPayload,
)
from app.services.product_categories import ProductCategoryService
from app.services.spus import SPUService
from app.services.skus import SKUService

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DATA_SHEET_NAME = "数据"
INSTRUCTION_SHEET_NAME = "填写说明"
REQUIRED_FONT_COLOR = "FFFF0000"


@dataclass(frozen=True)
class TemplateColumn:
    key: str
    title: str
    required: bool
    description: str


@dataclass(frozen=True)
class ImportTemplate:
    title: str
    columns: tuple[TemplateColumn, ...]
    notes: tuple[str, ...]


IMPORT_TEMPLATES: dict[str, ImportTemplate] = {
    "categories": ImportTemplate(
        title="分类导入模板",
        columns=(
            TemplateColumn("code", "分类编码", True, "全局唯一，长度不超过 50"),
            TemplateColumn("name", "分类名称", True, "分类显示名称"),
            TemplateColumn("level", "层级", True, "只能填写 1 / 2 / 3"),
            TemplateColumn("parent_code", "父级分类编码", False, "一级分类留空，二/三级分类必填"),
            TemplateColumn("sort_order", "排序", False, "非负整数，留空则按系统默认顺序"),
        ),
        notes=(
            "同一个文件内支持同时导入一级、二级、三级分类。",
            "二级分类的父级必须是一级分类，三级分类的父级必须是二级分类。",
            "若父级分类已在系统中存在，可直接填写父级编码；若父级也在本次导入中，请确保父级编码正确。",
        ),
    ),
    "spus": ImportTemplate(
        title="SPU导入模板",
        columns=(
            TemplateColumn("code", "SPU编码", True, "全局唯一，长度不超过 50"),
            TemplateColumn("name", "SPU名称", True, "长度不超过 100"),
            TemplateColumn("level1_category_code", "一级分类编码", True, "必须为已存在一级分类编码"),
            TemplateColumn("level2_category_code", "二级分类编码", True, "必须为已存在二级分类编码"),
            TemplateColumn("level3_category_code", "三级分类编码", True, "必须为已存在三级分类编码"),
            TemplateColumn("customer_warranty_months", "客户质保期(月)", True, "非负整数"),
            TemplateColumn("unit", "单位", True, "长度不超过 50"),
            TemplateColumn("restricted_countries", "禁止经营国家编码", False, "多个编码用逗号分隔，如 US,DE,GLOBAL"),
            TemplateColumn("supplier_name", "供应商名称", True, "长度不超过 100"),
            TemplateColumn("manufacturer_model", "厂家型号", True, "长度不超过 100"),
            TemplateColumn("purchase_price", "采购价", False, "非负数，最多两位小数"),
            TemplateColumn("purchase_warranty_months", "采购质保期(月)", False, "非负整数"),
            TemplateColumn("supplier_warranty_notes", "供应商质保说明", False, "长度不超过 500"),
            TemplateColumn("invoice_name", "开票品名", True, "同一 SPU 至少一条"),
            TemplateColumn("invoice_unit", "开票单位", True, "长度不超过 50"),
            TemplateColumn("invoice_model", "开票型号", True, "长度不超过 100"),
            TemplateColumn("company_subject", "公司抬头", True, "长度不超过 100"),
            TemplateColumn("invoice_sort_order", "开票排序", False, "非负整数，留空则按行顺序"),
        ),
        notes=(
            "一个 SPU 可占多行，每行表示一条开票信息；同一 SPU 的基础字段必须保持一致。",
            "分类必须预先存在于系统中，且层级关系必须匹配。",
            "禁止经营国家填写标准编码，多个值可用英文逗号、中文逗号或分号分隔。",
        ),
    ),
    "skus": ImportTemplate(
        title="SKU导入模板",
        columns=(
            TemplateColumn("spu_code", "所属SPU编码", True, "必须为已存在 SPU 编码"),
            TemplateColumn("code", "SKU编码", True, "全局唯一，长度不超过 50"),
            TemplateColumn("name_zh", "SKU中文名称", True, "长度不超过 100"),
            TemplateColumn("name_en", "SKU英文名称", True, "长度不超过 100"),
            TemplateColumn("product_model", "产品型号", True, "长度不超过 100"),
            TemplateColumn("product_type", "产品类型", True, "必须为启用的枚举 key"),
            TemplateColumn("core_params", "核心参数", True, "长度不超过 500"),
            TemplateColumn("product_status", "产品状态", False, "留空时默认上架，填写时必须为启用的枚举 key"),
            TemplateColumn("electrical_params", "电气参数", False, "长度不超过 100"),
            TemplateColumn("principle", "产品原理", True, "长度不超过 500"),
            TemplateColumn("usage", "用途", True, "长度不超过 500"),
            TemplateColumn("material", "材质", False, "长度不超过 200"),
            TemplateColumn("unit", "单位", True, "长度不超过 50"),
            TemplateColumn("has_plug", "是否带插头", True, "填写 true/false、是/否、1/0"),
            TemplateColumn("is_special", "是否特殊产品", True, "填写 true/false、是/否、1/0"),
            TemplateColumn("special_notes", "特殊说明", False, "长度不超过 1000"),
            TemplateColumn("package_type", "包装类型", False, "长度不超过 50"),
            TemplateColumn("package_quantity", "包装数量", False, "非负整数"),
            TemplateColumn("customs_hscode", "海关HS编码", False, "长度不超过 50"),
            TemplateColumn("customs_supervision_condition", "海关监管条件", False, "长度不超过 255"),
            TemplateColumn("customs_declaration_elements", "报关要素", False, "长度不超过 1000"),
            TemplateColumn("customs_refund_tax_rate", "退税率", False, "0-100，最多两位小数"),
            TemplateColumn("customs_info_ready", "报关信息已就绪", False, "填写 true/false、是/否、1/0"),
            TemplateColumn("package_net_weight_kg", "包装净重KG", False, "非负数，最多三位小数"),
            TemplateColumn("package_gross_weight_kg", "包装毛重KG", False, "非负数，最多三位小数"),
            TemplateColumn("package_length_cm", "包装长CM", False, "非负数，最多三位小数"),
            TemplateColumn("package_width_cm", "包装宽CM", False, "非负数，最多三位小数"),
            TemplateColumn("package_height_cm", "包装高CM", False, "非负数，最多三位小数"),
            TemplateColumn("package_volume_cbm", "包装体积CBM", False, "非负数，最多三位小数"),
            TemplateColumn("package_sort_order", "包装排序", False, "非负整数，留空则按行顺序"),
        ),
        notes=(
            "一个 SKU 可占多行，每行表示一条包装明细；同一 SKU 的基础字段必须保持一致。",
            "SKU 继承分类、供应商、禁止经营国家、客户质保期自所属 SPU，无需在模板中重复填写。",
            "若某个 SKU 不需要包装明细，可将包装明细相关列全部留空。",
        ),
    ),
}


class ImportTaskService:
    CATEGORY_TYPE = "categories"
    SPU_TYPE = "spus"
    SKU_TYPE = "skus"
    VALIDATED_STATUS = "validated"
    FAILED_VALIDATION_STATUS = "failed_validation"
    IMPORTED_STATUS = "imported"
    IMPORTING_STATUS = "importing"
    IMPORT_FAILED_STATUS = "import_failed"
    EXPIRED_HOURS = 24
    BOOLEAN_TRUE_VALUES = {"1", "true", "yes", "y", "是"}
    BOOLEAN_FALSE_VALUES = {"0", "false", "no", "n", "否"}
    COUNTRY_SPLIT_PATTERN = re.compile(r"[,，;；\n]+")

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ImportTaskRepository(db)
        self.category_repo = ProductCategoryRepository(db)
        self.spu_repo = SPURepository(db)
        self.sku_repo = SKURepository(db)

    async def download_template(self, import_type: str) -> tuple[str, bytes]:
        template = self._get_template(import_type)
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = DATA_SHEET_NAME
        headers = [column.title for column in template.columns]
        data_sheet.append(headers)
        for index, cell in enumerate(data_sheet[1]):
            column = template.columns[index]
            cell.font = Font(bold=True, color=REQUIRED_FONT_COLOR if column.required else None)

        instruction_sheet = workbook.create_sheet(INSTRUCTION_SHEET_NAME)
        instruction_sheet.append(["模板", template.title])
        instruction_sheet.append(["说明", "请勿修改数据页标题行；多行明细请保持基础字段一致。"])
        instruction_sheet.append([])
        instruction_sheet.append(["字段", "是否必填", "填写说明"])
        for cell in instruction_sheet[4]:
            cell.font = Font(bold=True)
        for column in template.columns:
            instruction_sheet.append([column.title, "是" if column.required else "否", column.description])
            row = instruction_sheet[instruction_sheet.max_row]
            if column.required:
                row[0].font = Font(color=REQUIRED_FONT_COLOR)
                row[1].font = Font(color=REQUIRED_FONT_COLOR)
        instruction_sheet.append([])
        instruction_sheet.append(["补充说明", ""])
        for note in template.notes:
            instruction_sheet.append(["", note])

        for sheet in (instruction_sheet, data_sheet):
            for index, column in enumerate(sheet.columns, start=1):
                max_length = max(len(str(cell.value or "")) for cell in column)
                sheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 4, 12), 36)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return f"{import_type}-import-template.xlsx", buffer.read()

    async def validate_import_file(
        self,
        import_type: str,
        file: UploadFile,
    ) -> ImportValidationResponse:
        await self.repo.soft_delete_expired()
        template = self._get_template(import_type)
        workbook = self._load_workbook(file)
        rows = self._read_data_rows(workbook, template)
        task = await self._build_validation_task(import_type, file.filename, rows)
        return self._serialize_validation_response(task)

    async def confirm_import(
        self,
        import_type: str,
        task_id: int,
        current_user: User,
    ) -> ImportConfirmResponse:
        task = await self._get_task_or_raise(task_id)
        self._ensure_task_type(task, import_type)
        self._ensure_not_expired(task)

        if task.status == self.IMPORTED_STATUS:
            imported_count = task.result_summary.get("imported_count", task.valid_rows) if task.result_summary else task.valid_rows
            return ImportConfirmResponse(
                task_id=task.id,
                task_type=task.task_type,
                status=task.status,
                imported_count=imported_count,
                progress_percent=task.progress_percent,
                confirmed_at=task.confirmed_at,
            )

        if task.status != self.VALIDATED_STATUS:
            raise BusinessError("当前导入任务未通过校验，无法确认导入")

        await self._persist_task_state(
            task.id,
            status=self.IMPORTING_STATUS,
            progress_percent=0,
            result_summary=None,
        )
        await self.db.refresh(task)

        try:
            imported_count = 0
            if import_type == self.CATEGORY_TYPE:
                imported_count = await self._confirm_categories(task, current_user)
            elif import_type == self.SPU_TYPE:
                imported_count = await self._confirm_spus(task, current_user)
            elif import_type == self.SKU_TYPE:
                imported_count = await self._confirm_skus(task, current_user)
            else:
                raise BusinessError("不支持的导入类型", code="NOT_FOUND", status_code=404)
        except Exception as exc:  # noqa: BLE001
            await self.db.rollback()
            await self._persist_task_state(
                task.id,
                status=self.IMPORT_FAILED_STATUS,
                result_summary={"error": self._describe_exception(exc)},
            )
            raise

        await self.db.refresh(task)
        task.status = self.IMPORTED_STATUS
        task.progress_percent = 100
        task.confirmed_at = datetime.now(timezone.utc)
        task.result_summary = {"imported_count": imported_count}
        self.db.add(task)
        await self.db.flush()

        return ImportConfirmResponse(
            task_id=task.id,
            task_type=task.task_type,
            status=task.status,
            imported_count=imported_count,
            progress_percent=task.progress_percent,
            confirmed_at=task.confirmed_at,
        )

    async def get_task_progress(self, task_id: int) -> ImportTaskProgressResponse:
        task = await self._get_task_or_raise(task_id)
        self._ensure_not_expired(task)
        return ImportTaskProgressResponse.model_validate(
            {
                "id": task.id,
                "task_type": task.task_type,
                "status": task.status,
                "original_filename": task.original_filename,
                "total_rows": task.total_rows,
                "valid_rows": task.valid_rows,
                "invalid_rows": task.invalid_rows,
                "progress_percent": task.progress_percent,
                "validation_errors": task.validation_errors,
                "result_summary": task.result_summary,
                "expires_at": task.expires_at,
                "confirmed_at": task.confirmed_at,
            }
        )

    def _load_workbook(self, file: UploadFile):
        filename = file.filename or ""
        if not filename.lower().endswith(".xlsx"):
            raise BusinessError("仅支持上传 .xlsx 文件")
        try:
            content = file.file.read()
            return load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise BusinessError("Excel 文件解析失败，请确认文件格式正确") from exc

    def _read_data_rows(
        self,
        workbook,
        template: ImportTemplate,
    ) -> list[dict[str, object]]:
        if DATA_SHEET_NAME not in workbook.sheetnames:
            raise BusinessError(f"Excel 模板缺少“{DATA_SHEET_NAME}”工作表")
        sheet = workbook[DATA_SHEET_NAME]
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            raise BusinessError("Excel 数据页为空")

        headers = [self._normalize_text(value) for value in raw_rows[0]]
        expected_headers = [column.title for column in template.columns]
        if headers != expected_headers:
            raise BusinessError("Excel 模板字段不匹配，请先下载最新模板后填写")

        rows: list[dict[str, object]] = []
        for row_number, values in enumerate(raw_rows[1:], start=2):
            if all(value in (None, "") for value in values):
                continue
            row = {column.key: values[index] for index, column in enumerate(template.columns)}
            row["_row_number"] = row_number
            rows.append(row)
        return rows

    async def _build_validation_task(
        self,
        import_type: str,
        filename: str | None,
        rows: list[dict[str, object]],
    ) -> ImportTask:
        if import_type == self.CATEGORY_TYPE:
            normalized_rows, errors = await self._validate_categories(rows)
        elif import_type == self.SPU_TYPE:
            normalized_rows, errors = await self._validate_spus(rows)
        elif import_type == self.SKU_TYPE:
            normalized_rows, errors = await self._validate_skus(rows)
        else:
            raise BusinessError("不支持的导入类型", code="NOT_FOUND", status_code=404)

        valid_rows = len(normalized_rows)
        invalid_rows = self._count_invalid_records(import_type, errors)
        task = ImportTask(
            task_type=import_type,
            status=self.VALIDATED_STATUS if invalid_rows == 0 else self.FAILED_VALIDATION_STATUS,
            original_filename=filename,
            total_rows=len(rows),
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            progress_percent=100,
            validation_errors=errors,
            payload=normalized_rows,
            result_summary=None,
            expires_at=datetime.now(timezone.utc) + timedelta(hours=self.EXPIRED_HOURS),
        )
        return await self.repo.save(task)

    async def _validate_categories(
        self,
        rows: list[dict[str, object]],
    ) -> tuple[list[dict], list[dict]]:
        errors: list[dict] = []
        if not rows:
            return [], []

        existing_categories = {item.code: item for item in await self.category_repo.get_all()}
        rows_by_code: dict[str, dict[str, object]] = {}
        level_by_code: dict[str, int] = {}
        normalized_rows: list[dict] = []

        for row in rows:
            row_number = int(row["_row_number"])
            code = self._required_string(row, "code", "分类编码", errors, row_number)
            name = self._required_string(row, "name", "分类名称", errors, row_number)
            level = self._required_int(row, "level", "层级", errors, row_number)
            parent_code = self._optional_string(row, "parent_code")
            sort_order = self._optional_int(row, "sort_order", "排序", errors, row_number)

            if code and code in existing_categories:
                self._add_error(errors, row_number, "分类编码", "分类编码已存在", code)
            if code and code in rows_by_code:
                self._add_error(errors, row_number, "分类编码", "同一文件中分类编码不可重复", code)
            if level is not None and level not in {1, 2, 3}:
                self._add_error(errors, row_number, "层级", "层级只能填写 1、2 或 3", code)
            if level == 1 and parent_code:
                self._add_error(errors, row_number, "父级分类编码", "一级分类不可填写父级编码", code)
            if level in {2, 3} and not parent_code:
                self._add_error(errors, row_number, "父级分类编码", "二级/三级分类必须填写父级编码", code)

            rows_by_code[code or f"__row__{row_number}"] = row
            if code and level is not None:
                level_by_code[code] = level
            normalized_rows.append(
                {
                    "row_number": row_number,
                    "code": code,
                    "name": name,
                    "level": level,
                    "parent_code": parent_code,
                    "sort_order": sort_order,
                }
            )

        for item in normalized_rows:
            code = item["code"]
            level = item["level"]
            parent_code = item["parent_code"]
            row_number = item["row_number"]
            if not code or level is None:
                continue
            if not parent_code:
                continue
            existing_parent = existing_categories.get(parent_code)
            imported_parent_level = level_by_code.get(parent_code)
            if existing_parent is None and imported_parent_level is None:
                self._add_error(errors, row_number, "父级分类编码", "父级分类编码不存在", code)
                continue
            parent_level = existing_parent.level if existing_parent is not None else imported_parent_level
            if parent_level != level - 1:
                self._add_error(errors, row_number, "父级分类编码", "父级分类层级不匹配", code)

        invalid_row_numbers = {item["row_number"] for item in errors}

        valid_rows: list[dict] = []
        normalized_rows.sort(key=lambda item: (item["level"], item["row_number"]))
        for item in normalized_rows:
            if item["row_number"] in invalid_row_numbers:
                continue
            payload = self._safe_model_validate(
                ProductCategoryCreate,
                {
                    "code": item["code"],
                    "name": item["name"],
                    "parent_id": None,
                    "sort_order": item["sort_order"],
                },
                errors=errors,
                row_number=item["row_number"],
                row_key=item["code"],
                field_labels={
                    "code": "分类编码",
                    "name": "分类名称",
                    "sort_order": "排序",
                },
            )
            if payload is not None:
                valid_rows.append(
                    {
                        "code": payload.code,
                        "name": payload.name,
                        "parent_code": item["parent_code"],
                        "sort_order": payload.sort_order,
                    }
                )
        return valid_rows, errors

    async def _validate_spus(
        self,
        rows: list[dict[str, object]],
    ) -> tuple[list[dict], list[dict]]:
        errors: list[dict] = []
        if not rows:
            return [], []

        categories = {item.code: item for item in await self.category_repo.get_all()}
        existing_spus = {item.code for item in await self.spu_repo.get_all()}
        existing_supplier_names = set(await self.spu_repo.list_supplier_names())

        grouped_rows: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in rows:
            row_number = int(row["_row_number"])
            code = self._required_string(row, "code", "SPU编码", errors, row_number)
            if code:
                if code in existing_spus:
                    self._add_error(errors, row_number, "SPU编码", "SPU编码已存在", code)
                grouped_rows[code].append(row)

        normalized_payloads: list[dict] = []
        for code, group_rows in grouped_rows.items():
            normalized = self._normalize_spu_group(
                code,
                group_rows,
                categories,
                existing_supplier_names,
                errors,
            )
            if normalized is not None:
                normalized_payloads.append(normalized)

        normalized_payloads.sort(key=lambda item: item["code"])
        return normalized_payloads, errors

    def _normalize_spu_group(
        self,
        code: str,
        rows: list[dict[str, object]],
        categories: dict[str, ProductCategory],
        existing_supplier_names: set[str],
        errors: list[dict],
    ) -> dict | None:
        local_errors: list[dict] = []
        first_row_number = int(rows[0]["_row_number"])
        base_fields = (
            "name",
            "level1_category_code",
            "level2_category_code",
            "level3_category_code",
            "customer_warranty_months",
            "unit",
            "restricted_countries",
            "supplier_name",
            "manufacturer_model",
            "purchase_price",
            "purchase_warranty_months",
            "supplier_warranty_notes",
        )
        base_values = {field: self._normalize_scalar(rows[0].get(field)) for field in base_fields}
        for row in rows[1:]:
            row_number = int(row["_row_number"])
            for field in base_fields:
                if self._normalize_scalar(row.get(field)) != base_values[field]:
                    self._add_error(local_errors, row_number, field, "同一 SPU 的基础字段必须保持一致", code)

        level1_code = self._required_string(rows[0], "level1_category_code", "一级分类编码", local_errors, first_row_number, code)
        level2_code = self._required_string(rows[0], "level2_category_code", "二级分类编码", local_errors, first_row_number, code)
        level3_code = self._required_string(rows[0], "level3_category_code", "三级分类编码", local_errors, first_row_number, code)
        category_ids = self._resolve_category_chain(level1_code, level2_code, level3_code, categories, local_errors, first_row_number, code)
        customer_warranty_months = self._required_int(rows[0], "customer_warranty_months", "客户质保期(月)", local_errors, first_row_number, code)
        purchase_warranty_months = self._optional_int(rows[0], "purchase_warranty_months", "采购质保期(月)", local_errors, first_row_number, code)
        purchase_price = self._optional_decimal_string(rows[0], "purchase_price", "采购价", local_errors, first_row_number, max_scale=2, row_key=code)
        restricted_countries = self._split_codes(rows[0].get("restricted_countries"))
        supplier_name = self._required_string(rows[0], "supplier_name", "供应商名称", local_errors, first_row_number, code)
        if supplier_name is not None and supplier_name not in existing_supplier_names:
            self._add_error(local_errors, first_row_number, "供应商名称", "供应商必须已存在", code)

        invoice_infos: list[dict] = []
        for index, row in enumerate(rows):
            row_number = int(row["_row_number"])
            invoice_name = self._required_string(row, "invoice_name", "开票品名", local_errors, row_number, code)
            invoice_unit = self._required_string(row, "invoice_unit", "开票单位", local_errors, row_number, code)
            invoice_model = self._required_string(row, "invoice_model", "开票型号", local_errors, row_number, code)
            company_subject = self._required_string(row, "company_subject", "公司抬头", local_errors, row_number, code)
            invoice_sort_order = self._optional_int(row, "invoice_sort_order", "开票排序", local_errors, row_number, code)
            if all(value is not None for value in (invoice_name, invoice_unit, invoice_model, company_subject)):
                invoice_infos.append(
                    {
                        "invoice_name": invoice_name,
                        "invoice_unit": invoice_unit,
                        "invoice_model": invoice_model,
                        "company_subject": company_subject,
                        "sort_order": invoice_sort_order if invoice_sort_order is not None else index,
                    }
                )

        if local_errors:
            errors.extend(local_errors)
            return None

        invoice_payloads: list[SPUInvoiceInfoPayload] = []
        for invoice_item in invoice_infos:
            payload = self._safe_model_validate(
                SPUInvoiceInfoPayload,
                invoice_item,
                errors=local_errors,
                row_number=rows[min(invoice_item["sort_order"], len(rows) - 1)]["_row_number"] if rows else first_row_number,
                row_key=code,
                field_labels={
                    "invoice_name": "开票品名",
                    "invoice_unit": "开票单位",
                    "invoice_model": "开票型号",
                    "company_subject": "公司抬头",
                    "sort_order": "开票排序",
                },
            )
            if payload is not None:
                invoice_payloads.append(payload)

        if local_errors:
            errors.extend(local_errors)
            return None

        payload = self._safe_model_validate(
            SPUCreate,
            {
                "code": code,
                "name": self._required_string(rows[0], "name", "SPU名称", local_errors, first_row_number, code),
                "level1_category_id": category_ids[0],
                "level2_category_id": category_ids[1],
                "level3_category_id": category_ids[2],
                "customer_warranty_months": customer_warranty_months,
                "unit": self._required_string(rows[0], "unit", "单位", local_errors, first_row_number, code),
                "restricted_countries": restricted_countries,
                "supplier_name": supplier_name,
                "manufacturer_model": self._required_string(rows[0], "manufacturer_model", "厂家型号", local_errors, first_row_number, code),
                "purchase_price": purchase_price,
                "purchase_warranty_months": purchase_warranty_months,
                "supplier_warranty_notes": self._optional_string(rows[0], "supplier_warranty_notes"),
                "invoice_infos": invoice_payloads,
            },
            errors=errors,
            row_number=first_row_number,
            row_key=code,
            field_labels={
                "code": "SPU编码",
                "name": "SPU名称",
                "level1_category_id": "一级分类编码",
                "level2_category_id": "二级分类编码",
                "level3_category_id": "三级分类编码",
                "customer_warranty_months": "客户质保期(月)",
                "unit": "单位",
                "restricted_countries": "禁止经营国家编码",
                "supplier_name": "供应商名称",
                "manufacturer_model": "厂家型号",
                "purchase_price": "采购价",
                "purchase_warranty_months": "采购质保期(月)",
                "supplier_warranty_notes": "供应商质保说明",
                "invoice_infos": "开票信息",
            },
        )
        return payload.model_dump(mode="json") if payload is not None else None

    async def _validate_skus(
        self,
        rows: list[dict[str, object]],
    ) -> tuple[list[dict], list[dict]]:
        errors: list[dict] = []
        if not rows:
            return [], []

        existing_spus = {item.code: item for item in await self.spu_repo.get_all()}
        existing_skus = {item.code for item in await self.sku_repo.get_all()}
        enums = await self.db.execute(
            select(SystemEnum).where(
                SystemEnum.deleted_at.is_(None),
                SystemEnum.is_enabled.is_(True),
                SystemEnum.enum_group.in_(["product_type", "product_status"]),
            )
        )
        enabled_enum_keys = defaultdict(set)
        for item in enums.scalars().all():
            enabled_enum_keys[item.enum_group].add(item.enum_key)

        grouped_rows: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in rows:
            row_number = int(row["_row_number"])
            code = self._required_string(row, "code", "SKU编码", errors, row_number)
            if code:
                if code in existing_skus:
                    self._add_error(errors, row_number, "SKU编码", "SKU编码已存在", code)
                grouped_rows[code].append(row)

        normalized_payloads: list[dict] = []
        for code, group_rows in grouped_rows.items():
            normalized = self._normalize_sku_group(code, group_rows, existing_spus, enabled_enum_keys, errors)
            if normalized is not None:
                normalized_payloads.append(normalized)

        normalized_payloads.sort(key=lambda item: item["create"]["code"])
        return normalized_payloads, errors

    def _normalize_sku_group(
        self,
        code: str,
        rows: list[dict[str, object]],
        spus: dict[str, SPU],
        enabled_enum_keys: dict[str, set[str]],
        errors: list[dict],
    ) -> dict | None:
        local_errors: list[dict] = []
        first_row_number = int(rows[0]["_row_number"])
        base_fields = (
            "spu_code",
            "name_zh",
            "name_en",
            "product_model",
            "product_type",
            "core_params",
            "product_status",
            "electrical_params",
            "principle",
            "usage",
            "material",
            "unit",
            "has_plug",
            "is_special",
            "special_notes",
            "package_type",
            "package_quantity",
            "customs_hscode",
            "customs_supervision_condition",
            "customs_declaration_elements",
            "customs_refund_tax_rate",
            "customs_info_ready",
        )
        base_values = {field: self._normalize_scalar(rows[0].get(field)) for field in base_fields}
        for row in rows[1:]:
            row_number = int(row["_row_number"])
            for field in base_fields:
                if self._normalize_scalar(row.get(field)) != base_values[field]:
                    self._add_error(local_errors, row_number, field, "同一 SKU 的基础字段必须保持一致", code)

        spu_code = self._required_string(rows[0], "spu_code", "所属SPU编码", local_errors, first_row_number, code)
        spu = spus.get(spu_code or "")
        if spu is None:
            self._add_error(local_errors, first_row_number, "所属SPU编码", "SPU编码不存在", code)

        product_type = self._required_string(rows[0], "product_type", "产品类型", local_errors, first_row_number, code)
        product_status = self._optional_string(rows[0], "product_status")
        if product_type and product_type not in enabled_enum_keys["product_type"]:
            self._add_error(local_errors, first_row_number, "产品类型", "产品类型必须为启用的枚举值", code)
        if product_status and product_status not in enabled_enum_keys["product_status"]:
            self._add_error(local_errors, first_row_number, "产品状态", "产品状态必须为启用的枚举值", code)

        has_plug = self._required_bool(rows[0], "has_plug", "是否带插头", local_errors, first_row_number, code)
        is_special = self._required_bool(rows[0], "is_special", "是否特殊产品", local_errors, first_row_number, code)
        package_quantity = self._optional_int(rows[0], "package_quantity", "包装数量", local_errors, first_row_number, code)
        customs_refund_tax_rate = self._optional_decimal_string(
            rows[0],
            "customs_refund_tax_rate",
            "退税率",
            local_errors,
            first_row_number,
            max_scale=2,
            row_key=code,
        )
        customs_info_ready = self._optional_bool(rows[0], "customs_info_ready", "报关信息已就绪", local_errors, first_row_number, code)

        package_details: list[dict] = []
        for index, row in enumerate(rows):
            row_number = int(row["_row_number"])
            package_detail = self._build_package_detail(row, row_number, local_errors, index, code)
            if package_detail is not None:
                package_details.append(package_detail)

        if local_errors:
            errors.extend(local_errors)
            return None

        package_payloads: list[SKUPackageDetailPayload] = []
        for detail in package_details:
            payload = self._safe_model_validate(
                SKUPackageDetailPayload,
                detail,
                errors=local_errors,
                row_number=first_row_number,
                row_key=code,
                field_labels={
                    "net_weight_kg": "包装净重KG",
                    "gross_weight_kg": "包装毛重KG",
                    "length_cm": "包装长CM",
                    "width_cm": "包装宽CM",
                    "height_cm": "包装高CM",
                    "volume_cbm": "包装体积CBM",
                    "sort_order": "包装排序",
                },
            )
            if payload is not None:
                package_payloads.append(payload)

        if local_errors:
            errors.extend(local_errors)
            return None

        create_payload = self._safe_model_validate(
            SKUCreate,
            {
                "spu_id": spu.id,
                "code": code,
                "name_zh": self._required_string(rows[0], "name_zh", "SKU中文名称", local_errors, first_row_number, code),
                "name_en": self._required_string(rows[0], "name_en", "SKU英文名称", local_errors, first_row_number, code),
                "product_model": self._required_string(rows[0], "product_model", "产品型号", local_errors, first_row_number, code),
                "product_type": product_type,
                "core_params": self._required_string(rows[0], "core_params", "核心参数", local_errors, first_row_number, code),
                "product_status": product_status,
                "electrical_params": self._optional_string(rows[0], "electrical_params"),
                "principle": self._required_string(rows[0], "principle", "产品原理", local_errors, first_row_number, code),
                "usage": self._required_string(rows[0], "usage", "用途", local_errors, first_row_number, code),
                "material": self._optional_string(rows[0], "material"),
                "unit": self._required_string(rows[0], "unit", "单位", local_errors, first_row_number, code),
                "has_plug": has_plug,
                "is_special": is_special,
                "special_notes": self._optional_string(rows[0], "special_notes"),
                "package_type": self._optional_string(rows[0], "package_type"),
                "package_quantity": package_quantity,
                "package_details": package_payloads,
            },
            errors=errors,
            row_number=first_row_number,
            row_key=code,
            field_labels={
                "spu_id": "所属SPU编码",
                "code": "SKU编码",
                "name_zh": "SKU中文名称",
                "name_en": "SKU英文名称",
                "product_model": "产品型号",
                "product_type": "产品类型",
                "core_params": "核心参数",
                "product_status": "产品状态",
                "electrical_params": "电气参数",
                "principle": "产品原理",
                "usage": "用途",
                "material": "材质",
                "unit": "单位",
                "has_plug": "是否带插头",
                "is_special": "是否特殊产品",
                "special_notes": "特殊说明",
                "package_type": "包装类型",
                "package_quantity": "包装数量",
                "package_details": "包装明细",
            },
        )
        customs_payload = self._safe_model_validate(
            SKUCustomsInfoUpdate,
            {
                "customs_hscode": self._optional_string(rows[0], "customs_hscode"),
                "customs_supervision_condition": self._optional_string(rows[0], "customs_supervision_condition"),
                "customs_declaration_elements": self._optional_string(rows[0], "customs_declaration_elements"),
                "customs_refund_tax_rate": customs_refund_tax_rate,
                "customs_info_ready": customs_info_ready,
            },
            errors=errors,
            row_number=first_row_number,
            row_key=code,
            field_labels={
                "customs_hscode": "海关HS编码",
                "customs_supervision_condition": "海关监管条件",
                "customs_declaration_elements": "报关要素",
                "customs_refund_tax_rate": "退税率",
                "customs_info_ready": "报关信息已就绪",
            },
        )
        if create_payload is None or customs_payload is None:
            return None
        return {
            "create": create_payload.model_dump(mode="json"),
            "customs": customs_payload.model_dump(mode="json"),
        }

    def _build_package_detail(
        self,
        row: dict[str, object],
        row_number: int,
        errors: list[dict],
        index: int,
        code: str,
    ) -> dict | None:
        field_map = {
            "net_weight_kg": ("package_net_weight_kg", "包装净重KG"),
            "gross_weight_kg": ("package_gross_weight_kg", "包装毛重KG"),
            "length_cm": ("package_length_cm", "包装长CM"),
            "width_cm": ("package_width_cm", "包装宽CM"),
            "height_cm": ("package_height_cm", "包装高CM"),
            "volume_cbm": ("package_volume_cbm", "包装体积CBM"),
        }
        values: dict[str, str | None] = {}
        has_any_value = False
        for payload_key, (row_key, label) in field_map.items():
            value = self._optional_decimal_string(row, row_key, label, errors, row_number, max_scale=3, row_key=code)
            values[payload_key] = value
            has_any_value = has_any_value or value is not None

        sort_order = self._optional_int(row, "package_sort_order", "包装排序", errors, row_number, code)
        if not has_any_value and sort_order is None:
            return None

        return {
            **values,
            "sort_order": sort_order if sort_order is not None else index,
        }

    async def _confirm_categories(self, task: ImportTask, current_user: User) -> int:
        service = ProductCategoryService(self.db)
        total = max(len(task.payload), 1)
        for index, item in enumerate(task.payload, start=1):
            parent_id = None
            parent_code = item.get("parent_code")
            if parent_code:
                parent = await self.category_repo.get_by_code(parent_code)
                if parent is None:
                    raise BusinessError(f"确认导入时父级分类不存在：{parent_code}")
                parent_id = parent.id
            payload = ProductCategoryCreate.model_validate(
                {
                    "code": item["code"],
                    "name": item["name"],
                    "parent_id": parent_id,
                    "sort_order": item.get("sort_order"),
                }
            )
            await service.create_category(payload)
            await self._persist_task_state(
                task.id,
                status=self.IMPORTING_STATUS,
                progress_percent=int(index * 100 / total),
            )
        return len(task.payload)

    async def _confirm_spus(self, task: ImportTask, current_user: User) -> int:
        service = SPUService(self.db)
        total = max(len(task.payload), 1)
        for index, item in enumerate(task.payload, start=1):
            payload = SPUCreate.model_validate(item)
            await service.create_spu(payload, current_user)
            await self._persist_task_state(
                task.id,
                status=self.IMPORTING_STATUS,
                progress_percent=int(index * 100 / total),
            )
        return len(task.payload)

    async def _confirm_skus(self, task: ImportTask, current_user: User) -> int:
        service = SKUService(self.db)
        total = max(len(task.payload), 1)
        for index, item in enumerate(task.payload, start=1):
            create_payload = SKUCreate.model_validate(item["create"])
            customs_payload = SKUCustomsInfoUpdate.model_validate(item["customs"])
            sku = await service.create_sku(create_payload, current_user)
            if customs_payload.model_dump(exclude_none=True):
                await service.update_customs_info(sku.id, customs_payload, current_user)
            await self._persist_task_state(
                task.id,
                status=self.IMPORTING_STATUS,
                progress_percent=int(index * 100 / total),
            )
        return len(task.payload)

    async def _get_task_or_raise(self, task_id: int) -> ImportTask:
        task = await self.repo.get_active_by_id(task_id)
        if task is None:
            raise BusinessError("导入任务不存在", code="NOT_FOUND", status_code=404)
        return task

    def _ensure_task_type(self, task: ImportTask, import_type: str) -> None:
        if task.task_type != import_type:
            raise BusinessError("导入任务类型不匹配")

    def _ensure_not_expired(self, task: ImportTask) -> None:
        expires_at = self._as_aware_datetime(task.expires_at)
        if expires_at is not None and expires_at < datetime.now(timezone.utc):
            raise BusinessError("导入任务已过期，请重新上传文件")

    def _as_aware_datetime(self, value: datetime | None) -> datetime | None:
        if value is None:
            return None
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value

    def _serialize_validation_response(self, task: ImportTask) -> ImportValidationResponse:
        return ImportValidationResponse(
            task_id=task.id,
            task_type=task.task_type,
            status=task.status,
            total_rows=task.total_rows,
            success_count=task.valid_rows,
            failed_count=task.invalid_rows,
            progress_percent=task.progress_percent,
            can_confirm=task.status == self.VALIDATED_STATUS,
            errors=[ImportValidationErrorItem.model_validate(item) for item in task.validation_errors],
        )

    def _get_template(self, import_type: str) -> ImportTemplate:
        template = IMPORT_TEMPLATES.get(import_type)
        if template is None:
            raise BusinessError("不支持的导入类型", code="NOT_FOUND", status_code=404)
        return template

    def _required_string(
        self,
        row: dict[str, object],
        key: str,
        field_label: str,
        errors: list[dict],
        row_number: int,
        row_key: str | None = None,
    ) -> str | None:
        value = self._optional_string(row, key)
        if value is None:
            self._add_error(errors, row_number, field_label, f"{field_label}不能为空", row_key)
        return value

    def _optional_string(self, row: dict[str, object], key: str) -> str | None:
        value = row.get(key)
        normalized = self._normalize_text(value)
        return normalized or None

    def _required_int(
        self,
        row: dict[str, object],
        key: str,
        field_label: str,
        errors: list[dict],
        row_number: int,
        row_key: str | None = None,
    ) -> int | None:
        value = self._optional_int(row, key, field_label, errors, row_number, row_key)
        if value is None:
            self._add_error(errors, row_number, field_label, f"{field_label}不能为空", row_key)
        return value

    def _optional_int(
        self,
        row: dict[str, object],
        key: str,
        field_label: str,
        errors: list[dict],
        row_number: int,
        row_key: str | None = None,
    ) -> int | None:
        raw = row.get(key)
        if raw in (None, ""):
            return None
        try:
            if isinstance(raw, bool):
                raise ValueError
            value = int(str(raw).strip())
        except (TypeError, ValueError):
            self._add_error(errors, row_number, field_label, f"{field_label}必须为整数", row_key)
            return None
        if value < 0:
            self._add_error(errors, row_number, field_label, f"{field_label}不能小于 0", row_key)
            return None
        return value

    def _required_bool(
        self,
        row: dict[str, object],
        key: str,
        field_label: str,
        errors: list[dict],
        row_number: int,
        row_key: str | None = None,
    ) -> bool | None:
        value = self._optional_bool(row, key, field_label, errors, row_number, row_key)
        if value is None:
            self._add_error(errors, row_number, field_label, f"{field_label}不能为空", row_key)
        return value

    def _optional_bool(
        self,
        row: dict[str, object],
        key: str,
        field_label: str,
        errors: list[dict],
        row_number: int,
        row_key: str | None = None,
    ) -> bool | None:
        raw = row.get(key)
        if raw in (None, ""):
            return None
        normalized = str(raw).strip().lower()
        if normalized in self.BOOLEAN_TRUE_VALUES:
            return True
        if normalized in self.BOOLEAN_FALSE_VALUES:
            return False
        self._add_error(errors, row_number, field_label, f"{field_label}必须填写 true/false 或 是/否", row_key)
        return None

    def _optional_decimal_string(
        self,
        row: dict[str, object],
        key: str,
        field_label: str,
        errors: list[dict],
        row_number: int,
        *,
        max_scale: int,
        row_key: str | None = None,
    ) -> str | None:
        raw = row.get(key)
        if raw in (None, ""):
            return None
        try:
            decimal_value = Decimal(str(raw).strip())
        except (InvalidOperation, ValueError):
            self._add_error(errors, row_number, field_label, f"{field_label}必须为数字", row_key)
            return None
        if decimal_value < 0:
            self._add_error(errors, row_number, field_label, f"{field_label}不能小于 0", row_key)
            return None
        quantized = format(decimal_value.normalize(), "f")
        if "." in quantized and len(quantized.split(".")[1]) > max_scale:
            self._add_error(errors, row_number, field_label, f"{field_label}最多保留 {max_scale} 位小数", row_key)
            return None
        return quantized

    def _count_invalid_records(self, import_type: str, errors: list[dict]) -> int:
        if not errors:
            return 0
        if import_type == self.CATEGORY_TYPE:
            return len({item["row_number"] for item in errors})
        return len({item.get("row_key") or f"row:{item['row_number']}" for item in errors})

    def _describe_exception(self, exc: Exception) -> str:
        if isinstance(exc, BusinessError):
            return exc.message
        return "导入执行失败"

    def _safe_model_validate(
        self,
        model_cls,
        payload: dict,
        *,
        errors: list[dict],
        row_number: int,
        row_key: str | None = None,
        field_labels: dict[str, str] | None = None,
    ):
        try:
            return model_cls.model_validate(payload)
        except ValidationError as exc:
            for item in exc.errors():
                loc = item.get("loc") or ()
                field = str(loc[-1]) if loc else "数据"
                display_field = field_labels.get(field, field) if field_labels else field
                self._add_error(errors, row_number, display_field, item["msg"], row_key)
            return None

    async def _persist_task_state(self, task_id: int, **updates) -> None:
        session_factory = async_sessionmaker(self.db.bind, expire_on_commit=False)
        async with session_factory() as session:
            task = await session.scalar(
                select(ImportTask).where(
                    ImportTask.id == task_id,
                    ImportTask.deleted_at.is_(None),
                )
            )
            if task is None:
                return
            for field, value in updates.items():
                setattr(task, field, value)
            session.add(task)
            await session.commit()

    def _resolve_category_chain(
        self,
        level1_code: str | None,
        level2_code: str | None,
        level3_code: str | None,
        categories: dict[str, ProductCategory],
        errors: list[dict],
        row_number: int,
        row_key: str,
    ) -> tuple[int, int, int]:
        level1 = categories.get(level1_code or "")
        level2 = categories.get(level2_code or "")
        level3 = categories.get(level3_code or "")
        if level1 is None:
            self._add_error(errors, row_number, "一级分类编码", "一级分类编码不存在", row_key)
        if level2 is None:
            self._add_error(errors, row_number, "二级分类编码", "二级分类编码不存在", row_key)
        if level3 is None:
            self._add_error(errors, row_number, "三级分类编码", "三级分类编码不存在", row_key)
        if level1 is not None and level1.level != 1:
            self._add_error(errors, row_number, "一级分类编码", "一级分类编码必须指向一级分类", row_key)
        if level2 is not None and level2.level != 2:
            self._add_error(errors, row_number, "二级分类编码", "二级分类编码必须指向二级分类", row_key)
        if level3 is not None and level3.level != 3:
            self._add_error(errors, row_number, "三级分类编码", "三级分类编码必须指向三级分类", row_key)
        if level1 is not None and level2 is not None and level2.parent_id != level1.id:
            self._add_error(errors, row_number, "二级分类编码", "分类层级关系不匹配", row_key)
        if level2 is not None and level3 is not None and level3.parent_id != level2.id:
            self._add_error(errors, row_number, "三级分类编码", "分类层级关系不匹配", row_key)
        return (
            level1.id if level1 is not None else 0,
            level2.id if level2 is not None else 0,
            level3.id if level3 is not None else 0,
        )

    def _split_codes(self, raw: object) -> list[str]:
        value = self._normalize_text(raw)
        if not value:
            return []
        result: list[str] = []
        seen: set[str] = set()
        for part in self.COUNTRY_SPLIT_PATTERN.split(value):
            normalized = part.strip().upper()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            result.append(normalized)
        return result

    def _normalize_text(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _normalize_scalar(self, value: object) -> object:
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip()
        return value

    def _add_error(
        self,
        errors: list[dict],
        row_number: int,
        field: str,
        message: str,
        row_key: str | None = None,
    ) -> None:
        errors.append(
            {
                "row_number": row_number,
                "field": field,
                "message": message,
                "row_key": row_key,
            }
        )
